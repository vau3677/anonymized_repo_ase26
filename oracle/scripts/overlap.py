#!/usr/bin/env python3
"""
TOSEM -> Web3Bugs ID -> MV-Bench matcher
1. Read TOSEM seed rows from oracle/tosem_seed_raw.csv.
2. For each TOSEM row, split finding_slug into:
     repo_slug   = text before '#', e.g. 2021-07-wildcredit
     finding_id  = text after '#',  e.g. h-02
3. Find Web3Bugs ID XX by searching Web3Bugs/reports/XX.md for:
     slug: <repo_slug>
4. Fallback: search Web3Bugs/contracts/XX/README.md for <repo_slug>.
5. Find MV-Bench sheets whose names start with XX, with the next char not a digit.
     XX=1  matches 1, 1contracts
     XX=1  does not match 10contracts, 112backd
     XX=18 matches 18, 18contracts
6. In each candidate sheet, output every row containing the same finding_id.
7. Output TOSEM row and full MV-Bench row side-by-side.
"""
from __future__ import annotations
import argparse, csv, json, re
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

JOIN_COLUMNS = [
    # TOSEM fields
    "oracle_id",
    "source_dataset",
    "source_category",
    "source_subcategory",
    "serial_number",
    "reporting_time",
    "bug_link",
    "finding_slug",
    "repo_slug",
    "finding_id_raw",
    "finding_id_norm",
    "source_root_cause",
    "source_exploitation_method",
    "source_fix_strategy",
    "raw_line",
    "source_file",

    # Web3Bugs bridge
    "web3bugs_status",
    "web3bugs_id",
    "web3bugs_dir",
    "web3bugs_source_file",
    "web3bugs_match_basis",
    "web3bugs_match_line",

    # MV-Bench sheet / row
    "mvbench_status",
    "mvbench_candidate_sheets",
    "matched_sheet",
    "mvbench_row_idx",
    "mvbench_ablation",
    "mvbench_type",
    "mvbench_code4rena_report",
    "mvbench_comments",
    "mvbench_finding_header",
    "mvbench_finding",
    "mvbench_row_text",
    "mvbench_row_json",
]

SUMMARY_COLUMNS = [
    "oracle_id",
    "finding_slug",
    "repo_slug",
    "finding_id_norm",
    "source_category",
    "web3bugs_status",
    "web3bugs_ids",
    "web3bugs_dirs",
    "web3bugs_basis",
    "mvbench_status",
    "mvbench_candidate_sheets",
    "finding_match_count",
    "b0_match_count",
    "matched_ablations",
    "matched_types",
    "source_root_cause",
    "source_exploitation_method",
]

def clean(s): return re.sub(r"\s+", " ", (s or "").replace("\u2028", "\n").replace("\xa0", " ")).strip()

def slug_from_bug_link(url: str) -> str:
    m = re.search(r"/reports/([^#/?]+)#([hmql]-\d+)", url or "", re.I)
    if not m: return ""
    return f"{m.group(1).lower()}#{m.group(2).lower()}"

def split_finding_slug(slug: str) -> tuple[str, str]:
    slug = (slug or "").strip()
    if "#" not in slug: return slug.lower(), ""
    repo, finding = slug.split("#", 1)
    return repo.strip().lower(), finding.strip().lower()

# h-2, h-02, H_02, H 02 -> H-02
def normalize_finding_id(fid: str):
    m = re.match(r"^\s*([hmql])[-_ ]?0*(\d+)\s*$", fid or "", re.I)
    if not m: return (fid or "").strip().upper()
    return f"{m.group(1).upper()}-{int(m.group(2)):02d}"

# H-02 matches: H-02, h-02, H-2, #h-02, h_02, h 02
def finding_regex(fid_norm: str) -> re.Pattern:
    m = re.match(r"^([HMQL])-(\d+)$", fid_norm)
    if not m: return re.compile(re.escape(fid_norm), re.I)
    letter = m.group(1)
    number = int(m.group(2))
    return re.compile(rf"(?<![A-Za-z0-9])#?{letter}[-_ ]?0*{number}(?![0-9])", re.I)

def load_tosem_seed(path: Path) -> list[dict[str, str]]:
    if not path.exists(): raise FileNotFoundError(f"TOSEM seed not found: {path}")

    rows = []
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            finding_slug = row.get("finding_slug", "").strip().lower()

            if "#" not in finding_slug:
                finding_slug = slug_from_bug_link(row.get("bug_link", ""))
            repo_slug, finding_id = split_finding_slug(finding_slug)
            row["finding_slug"] = finding_slug
            row["repo_slug"] = repo_slug
            row["finding_id_raw"] = finding_id
            row["finding_id_norm"] = normalize_finding_id(finding_id)
            rows.append(row)
    return rows

def read_text(path: Path) -> str: return path.read_text(encoding="utf-8", errors="replace")

def extract_id_from_report_file(path: Path) -> str:
    m = re.match(r"^(\d+)\.md$", path.name)
    return m.group(1) if m else ""

def extract_id_from_contract_readme(path: Path) -> str:
    m = re.match(r"^(\d+)", path.parent.name)
    return m.group(1) if m else ""

def web3bugs_reports(web3bugs_root: Path) -> list[Path]:
    reports_dir = web3bugs_root / "reports"
    if not reports_dir.exists(): return []
    return sorted(reports_dir.glob("*.md"))

def web3bugs_contract_readmes(contracts_dir: Path) -> list[Path]:
    if not contracts_dir.exists(): raise FileNotFoundError(f"Web3Bugs contracts dir not found: {contracts_dir}")
    return sorted(contracts_dir.glob("*/README.md"))

def find_by_report_slug(repo_slug: str, web3bugs_root: Path) -> list[dict[str, str]]:
    repo_slug = (repo_slug or "").strip().lower()
    if not repo_slug: return []
    matches = []
    for report in web3bugs_reports(web3bugs_root):
        text = read_text(report)

        for line in text.splitlines():
            m = re.match(r"^\s*slug\s*:\s*(\S+)\s*$", line.strip(), re.I)
            if not m: continue

            slug_value = m.group(1).strip().lower()
            if slug_value != repo_slug: continue

            wid = extract_id_from_report_file(report)
            if not wid: continue

            matches.append({
                "web3bugs_id": wid,
                "web3bugs_dir": wid,
                "web3bugs_source_file": str(report),
                "web3bugs_match_basis": "reports_slug",
                "web3bugs_match_line": line.strip(),
            })

    return sorted(matches, key=lambda r: int(r["web3bugs_id"]))

def find_by_report_text(repo_slug: str, web3bugs_root: Path) -> list[dict[str, str]]:
    repo_slug = (repo_slug or "").strip().lower()
    if not repo_slug: return []

    matches = []
    for report in web3bugs_reports(web3bugs_root):
        text = read_text(report)
        text_lower = text.lower()

        if repo_slug not in text_lower: continue

        wid = extract_id_from_report_file(report)
        if not wid: continue

        hit_line = ""
        for line in text.splitlines():
            if repo_slug in line.lower():
                hit_line = clean(line)
                break

        matches.append({
            "web3bugs_id": wid,
            "web3bugs_dir": wid,
            "web3bugs_source_file": str(report),
            "web3bugs_match_basis": "reports_text",
            "web3bugs_match_line": hit_line,
        })

    return sorted(matches, key=lambda r: int(r["web3bugs_id"]))

def find_by_contract_readme(repo_slug: str, readmes: list[Path]) -> list[dict[str, str]]:
    repo_slug = (repo_slug or "").strip().lower()
    if not repo_slug: return []

    matches = []
    for readme in readmes:
        text = read_text(readme)
        text_lower = text.lower()

        if repo_slug not in text_lower: continue

        wid = extract_id_from_contract_readme(readme)
        if not wid: continue

        hit_line = ""
        for line in text.splitlines():
            if repo_slug in line.lower():
                hit_line = clean(line)
                break

        matches.append({
            "web3bugs_id": wid,
            "web3bugs_dir": readme.parent.name,
            "web3bugs_source_file": str(readme),
            "web3bugs_match_basis": "contracts_readme",
            "web3bugs_match_line": hit_line,
        })
    return sorted(matches, key=lambda r: int(r["web3bugs_id"]))

def dedupe_web3bugs_matches(matches: list[dict[str, str]]) -> list[dict[str, str]]:
    by_id = {}
    for m in matches:
        wid = m.get("web3bugs_id", "")
        if not wid: continue
        if wid not in by_id: by_id[wid] = m
    return [by_id[k] for k in sorted(by_id, key=int)]

# Lookup order: reports/XX.md exact slug field, contracts/XX/README.md fallback
def find_web3bugs_ids(repo_slug, web3bugs_root, contract_readmes):
    hits = find_by_report_slug(repo_slug, web3bugs_root)
    if hits: return dedupe_web3bugs_matches(hits)
    hits = find_by_contract_readme(repo_slug, contract_readmes)
    return dedupe_web3bugs_matches(hits)

def cell_text(cell) -> str:
    parts = []
    if cell.value is not None: parts.append(str(cell.value))

    try:
        if cell.hyperlink and cell.hyperlink.target:
            parts.append(str(cell.hyperlink.target))
    except Exception:
        pass
    return clean(" ".join(parts))

def load_mvbench(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"MV-Bench workbook not found: {path}")

    wb = load_workbook(path, data_only=True, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        if ws.title == "B0_FILL_LOG": continue

        parsed_rows = []
        for row in ws.iter_rows():
            values = [cell_text(c) for c in row]
            if not any(values): continue
            parsed_rows.append({ "row_idx": row[0].row, "values": values })

        headers = parsed_rows[0]["values"] if parsed_rows else []
        sheets.append({ "sheet_name": ws.title, "headers": headers, "rows": parsed_rows })
    return sheets

# ID=1 matches "1", "1contracts"
def sheet_matches_web3bugs_id(sheet_name: str, web3bugs_id: str): return re.match( rf"^{re.escape(web3bugs_id)}($|[^0-9])", sheet_name or "") is not None

def row_to_dict(headers: list[str], values: list[str]) -> dict[str, str]:
    out = {}
    for i, value in enumerate(values):
        if i < len(headers) and headers[i]: key = headers[i]
        else: key = f"col_{i + 1}"
        out[key] = value
    return out

def get_field(row: dict[str, str], *names: str) -> str:
    lower = {k.lower(): v for k, v in row.items()}
    for name in names:
        if name.lower() in lower: return lower[name.lower()]
    return ""

def row_text(row): return "\n".join(v for v in row.values() if v)
def row_matches_finding(row, fid_norm): return finding_regex(fid_norm).search(row_text(row)) is not None
def finding_header(finding: str) -> str: return (finding or "").split("\n", 1)[0].strip()

def tosem_base(row: dict[str, str]) -> dict[str, str]:
    return {
        "oracle_id": row.get("oracle_id", ""),
        "source_dataset": row.get("source_dataset", ""),
        "source_category": row.get("source_category", ""),
        "source_subcategory": row.get("source_subcategory", ""),
        "serial_number": row.get("serial_number", ""),
        "reporting_time": row.get("reporting_time", ""),
        "bug_link": row.get("bug_link", ""),
        "finding_slug": row.get("finding_slug", ""),
        "repo_slug": row.get("repo_slug", ""),
        "finding_id_raw": row.get("finding_id_raw", ""),
        "finding_id_norm": row.get("finding_id_norm", ""),
        "source_root_cause": row.get("source_root_cause", ""),
        "source_exploitation_method": row.get("source_exploitation_method", ""),
        "source_fix_strategy": row.get("source_fix_strategy", ""),
        "raw_line": row.get("raw_line", ""),
        "source_file": row.get("source_file", ""),
    }

def write_csv(path: Path, rows, columns):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows: writer.writerow({c: row.get(c, "") for c in columns})

def match_all(tosem_rows, web3bugs_root, contract_readmes, sheets):
    join_rows, summary_rows = [], []
    for case in tosem_rows:
        repo_slug = case["repo_slug"]
        fid_norm = case["finding_id_norm"]
        web3_matches = find_web3bugs_ids(repo_slug, web3bugs_root, contract_readmes)
        web3_ids = sorted({m["web3bugs_id"] for m in web3_matches}, key=int)

        candidate_sheets = [sheet for sheet in sheets if any(sheet_matches_web3bugs_id(sheet["sheet_name"], wid) for wid in web3_ids)]
        matched_rows = []
        for sheet in candidate_sheets:
            headers = sheet["headers"]
            for parsed in sheet["rows"][1:]:
                row_dict = row_to_dict(headers, parsed["values"])
                if row_matches_finding(row_dict, fid_norm):
                    matched_rows.append((sheet, parsed["row_idx"], row_dict))

        web3_status = "ambiguous_multiple_web3bugs_ids"
        if not web3_ids: web3_status = "no_web3bugs_id"
        elif len(web3_ids) == 1: web3_status = "web3bugs_id_found"

        mvbench_status = "no_web3bugs_id"
        if matched_rows: mvbench_status = "finding_match"
        elif candidate_sheets: mvbench_status = "repo_sheet_no_finding"
        elif web3_ids: mvbench_status = "web3bugs_id_no_mvbench_sheet"

        web3_id_text = ";".join(web3_ids)
        web3_dir_text = ";".join(sorted({m["web3bugs_dir"] for m in web3_matches}))
        web3_source_text = ";".join(sorted({m["web3bugs_source_file"] for m in web3_matches}))
        web3_basis_text = ";".join(sorted({m["web3bugs_match_basis"] for m in web3_matches}))
        web3_line_text = " || ".join(
            m["web3bugs_match_line"]
            for m in web3_matches
            if m.get("web3bugs_match_line")
        )
        candidate_sheet_names = ";".join(s["sheet_name"] for s in candidate_sheets)

        if matched_rows:
            for sheet, row_idx, row_dict in matched_rows:
                finding = get_field(row_dict, "Finding")

                out = tosem_base(case)
                out.update({
                    "web3bugs_status": web3_status,
                    "web3bugs_id": web3_id_text,
                    "web3bugs_dir": web3_dir_text,
                    "web3bugs_source_file": web3_source_text,
                    "web3bugs_match_basis": web3_basis_text,
                    "web3bugs_match_line": web3_line_text,
                    "mvbench_status": mvbench_status,
                    "mvbench_candidate_sheets": candidate_sheet_names,
                    "matched_sheet": sheet["sheet_name"],
                    "mvbench_row_idx": str(row_idx),
                    "mvbench_ablation": get_field(row_dict, "Ablation"),
                    "mvbench_type": get_field(row_dict, "Type"),
                    "mvbench_code4rena_report": get_field(row_dict, "Code4rena Report"),
                    "mvbench_comments": get_field(row_dict, "Comments"),
                    "mvbench_finding_header": finding_header(finding),
                    "mvbench_finding": finding,
                    "mvbench_row_text": row_text(row_dict),
                    "mvbench_row_json": json.dumps(row_dict, ensure_ascii=False),
                })
                join_rows.append(out)
        else:
            out = tosem_base(case)
            out.update({
                "web3bugs_status": web3_status,
                "web3bugs_id": web3_id_text,
                "web3bugs_dir": web3_dir_text,
                "web3bugs_source_file": web3_source_text,
                "web3bugs_match_basis": web3_basis_text,
                "web3bugs_match_line": web3_line_text,
                "mvbench_status": mvbench_status,
                "mvbench_candidate_sheets": candidate_sheet_names,
            })
            join_rows.append(out)

        ablations = Counter(get_field(row_dict, "Ablation") or "<blank>" for _, _, row_dict in matched_rows)
        types = Counter(get_field(row_dict, "Type") or "<blank>" for _, _, row_dict in matched_rows)

        summary_rows.append({
            "oracle_id": case.get("oracle_id", ""),
            "finding_slug": case.get("finding_slug", ""),
            "repo_slug": repo_slug,
            "finding_id_norm": fid_norm,
            "source_category": case.get("source_category", ""),
            "web3bugs_status": web3_status,
            "web3bugs_ids": web3_id_text,
            "web3bugs_dirs": web3_dir_text,
            "web3bugs_basis": web3_basis_text,
            "mvbench_status": mvbench_status,
            "mvbench_candidate_sheets": candidate_sheet_names,
            "finding_match_count": str(len(matched_rows)),
            "b0_match_count": str(sum(1 for _, _, row_dict in matched_rows if get_field(row_dict, "Ablation").strip() == "B0")),
            "matched_ablations": json.dumps(dict(ablations), ensure_ascii=False),
            "matched_types": json.dumps(dict(types), ensure_ascii=False),
            "source_root_cause": case.get("source_root_cause", ""),
            "source_exploitation_method": case.get("source_exploitation_method", ""),
        })

    return join_rows, summary_rows

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--tosem-seed", default="oracle/tosem_seed_raw.csv")
    parser.add_argument("--web3bugs-contracts", default="dataset/Web3Bugs/contracts")
    parser.add_argument("--mvbench", default="MV-Bench.xlsx")
    parser.add_argument("--out", default="oracle/tosem_web3bugs_mvbench_join.csv")
    parser.add_argument("--summary", default="oracle/tosem_web3bugs_mvbench_summary.csv")
    args = parser.parse_args()

    contracts_dir = Path(args.web3bugs_contracts)
    web3bugs_root = contracts_dir.parent

    tosem_rows = load_tosem_seed(Path(args.tosem_seed))
    contract_readmes = web3bugs_contract_readmes(contracts_dir)
    sheets = load_mvbench(Path(args.mvbench))
    join_rows, summary_rows = match_all(tosem_rows, web3bugs_root, contract_readmes, sheets)
    write_csv(Path(args.out), join_rows, JOIN_COLUMNS)
    write_csv(Path(args.summary), summary_rows, SUMMARY_COLUMNS)

    print(f"[OK] TOSEM rows: {len(tosem_rows)}")
    print(f"[OK] Web3Bugs contract READMEs: {len(contract_readmes)}")
    print(f"[OK] MV-Bench sheets: {len(sheets)}")
    print(f"[OK] Wrote {args.out}")
    print(f"[OK] Wrote {args.summary}")
    print("[SUMMARY] Web3Bugs:", dict(Counter(row["web3bugs_status"] for row in summary_rows)))
    print("[SUMMARY] MV-Bench:", dict(Counter(row["mvbench_status"] for row in summary_rows)))
    print("[SUMMARY] Total MV-Bench matched rows:", sum(int(row["finding_match_count"]) for row in summary_rows))
    print("[SUMMARY] Total B0 matched rows:", sum(int(row["b0_match_count"]) for row in summary_rows))

if __name__ == "__main__": main()